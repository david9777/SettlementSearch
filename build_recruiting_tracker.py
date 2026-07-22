# -*- coding: utf-8 -*-
"""Build the Recruiting Tracker workbook: Tracker, Dashboard, Lists."""
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

OUT = r"C:\Users\DSamson\.claude\Settlement\Recruiting_Tracker.xlsx"
LAST = 500  # last row that gets dropdowns / formulas / formatting

# ---- dropdown value lists -------------------------------------------------
ROLES = ["Paralegal", "Legal Assistant", "Attorney – Mass Arbitration",
         "Attorney – Data Breach", "Attorney – Privacy",
         "Partner-Level Associate Attorney"]
SOURCES = ["Referral", "LinkedIn (Active)", "LinkedIn (Passive)",
           "Job Posting / Applicant", "Internal", "University / School", "Other"]
METHODS = ["LinkedIn Active", "LinkedIn Passive", "Referral", "Email", "Other"]
YESNO = ["Yes", "No"]
STATUS = ["Not Scheduled", "Scheduled", "Completed", "Passed", "Rejected",
          "On Hold", "Cancelled"]
FR_INT = ["Amanda", "Tyler", "Melissa", "Michael", "Trent", "Eleanor", "Chris",
          "Gary", "Bhagavat"]
PARTNER = ["Marc Reich", "Ed Korsinsky"]
REFSTATUS = ["Not Started", "In Progress", "Completed"]
OUTCOME = ["Accepted", "Declined", "Pending"]

# ---- fonts / fills --------------------------------------------------------
F_BASE = Font(name="Calibri", size=11)
F_BOLD = Font(name="Calibri", size=11, bold=True)
F_HEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_AUTO = Font(name="Calibri", size=11, italic=True, color="3F3F3F")
F_TITLE = Font(name="Calibri", size=16, bold=True, color="1F4E78")
F_SUB = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
F_NOTE = Font(name="Calibri", size=9, italic=True, color="808080")

FILL_HEAD = PatternFill("solid", fgColor="1F4E78")
FILL_AUTOHEAD = PatternFill("solid", fgColor="375623")
FILL_SUB = PatternFill("solid", fgColor="2E75B6")
FILL_TOTAL = PatternFill("solid", fgColor="DDEBF7")
FILL_LISTHEAD = PatternFill("solid", fgColor="808080")

ROLE_FILLS = {
    "Paralegal": "BDD7EE",
    "Legal Assistant": "C6E0B4",
    "Attorney – Mass Arbitration": "FFE699",
    "Attorney – Data Breach": "F4B183",
    "Attorney – Privacy": "D5A6BD",
    "Partner-Level Associate Attorney": "B4A7D6",
}

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ==========================================================================
# LISTS SHEET
# ==========================================================================
ls = wb.active
ls.title = "Lists"
list_cols = [
    ("Role / Position", ROLES),
    ("Source", SOURCES),
    ("Contact Method", METHODS),
    ("Yes / No", YESNO),
    ("Interview Status", STATUS),
    ("First Round Interviewer", FR_INT),
    ("Partner Interviewer", PARTNER),
    ("Reference Check Status", REFSTATUS),
    ("Offer Outcome", OUTCOME),
]
for ci, (hdr, vals) in enumerate(list_cols, start=1):
    c = ls.cell(row=1, column=ci, value=hdr)
    c.font = F_HEAD
    c.fill = FILL_LISTHEAD
    c.alignment = center
    ls.column_dimensions[get_column_letter(ci)].width = max(16, len(hdr) + 2)
    for ri, v in enumerate(vals, start=2):
        cc = ls.cell(row=ri, column=ci, value=v)
        cc.font = F_BASE
ls.cell(row=len(max(list_cols, key=lambda t: len(t[1]))[1]) + 3, column=1,
        value="Reference lists that drive the dropdowns on the Tracker tab. "
              "Edit values here to change the available options.").font = F_NOTE

# range strings for data validation (cross-sheet references)
def rng(col_idx, n):
    L = get_column_letter(col_idx)
    return f"Lists!${L}$2:${L}${1 + n}"

VR = {
    "role": rng(1, len(ROLES)),
    "source": rng(2, len(SOURCES)),
    "method": rng(3, len(METHODS)),
    "yesno": rng(4, len(YESNO)),
    "status": rng(5, len(STATUS)),
    "frint": rng(6, len(FR_INT)),
    "partner": rng(7, len(PARTNER)),
    "ref": rng(8, len(REFSTATUS)),
    "outcome": rng(9, len(OUTCOME)),
}

# ==========================================================================
# TRACKER SHEET
# ==========================================================================
tr = wb.create_sheet("Tracker", 0)

HEADERS = [
    ("Candidate Name", 22, None),
    ("Role / Position", 26, "role"),
    ("Date Contacted", 14, "date"),
    ("Source", 20, "source"),
    ("Contact Method", 16, "method"),
    ("Screened?", 11, "yesno"),
    ("Screened Date", 14, "date"),
    ("First Round Interview Status", 20, "status"),
    ("First Round Interviewer", 17, "frint"),
    ("First Round Interview Date", 15, "date"),
    ("Second Round Status (Partner Round)", 20, "status"),
    ("Partner Interviewer", 16, "partner"),
    ("Partner Interview Date", 15, "date"),
    ("Assessment Required?", 13, "yesno"),
    ("Assessment Sent Date", 14, "date"),
    ("Assessment Returned?", 13, "yesno"),
    ("Assessment Return Date", 14, "date"),
    ("Reference Check Status", 16, "ref"),
    ("Offer Initiated?", 12, "yesno"),
    ("Offer Initiated Date", 14, "date"),
    ("Offer Outcome", 13, "outcome"),
    ("Final Status Notes", 34, None),
    ("Pipeline Stage", 14, "auto"),
    ("Active / Closed", 13, "auto"),
    ("Days in Pipeline", 12, "auto"),
]
NCOL = len(HEADERS)  # 25 -> A..Y

# header row
tr.row_dimensions[1].height = 42
for ci, (name, width, kind) in enumerate(HEADERS, start=1):
    c = tr.cell(row=1, column=ci, value=name)
    c.font = F_AUTO if kind == "auto" else F_HEAD
    c.fill = FILL_AUTOHEAD if kind == "auto" else FILL_HEAD
    c.alignment = center
    c.border = BORDER
    tr.column_dimensions[get_column_letter(ci)].width = width

# date columns -> number format applied later per row
DATE_COLS = [ci for ci, (_, _, k) in enumerate(HEADERS, start=1) if k == "date"]

# auto formula columns: W=23 Pipeline Stage, X=24 Active/Closed, Y=25 Days
def pipeline_formula(r):
    return (
        f'=IF($A{r}="","",'
        f'IF($U{r}="Accepted","Hired",'
        f'IF($S{r}="Yes","Offer",'
        f'IF(OR($R{r}="In Progress",$R{r}="Completed"),"Reference",'
        f'IF(OR($P{r}="Yes",$O{r}<>""),"Assessment",'
        f'IF(OR($M{r}<>"",$L{r}<>"",$K{r}<>""),"Partner",'
        f'IF(OR($J{r}<>"",$I{r}<>"",$H{r}<>""),"First Round",'
        f'IF($F{r}="Yes","Screened","Contacted"))))))))'
    )

def status_formula(r):
    return (
        f'=IF($A{r}="","",'
        f'IF(OR($U{r}="Accepted",$U{r}="Declined",$H{r}="Rejected",'
        f'$K{r}="Rejected"),"Closed","Active"))'
    )

def days_formula(r):
    return f'=IF($C{r}="","",TODAY()-$C{r})'

# ---- sample candidates ----------------------------------------------------
samples = [
    # A name, B role, C contacted, D source, E method, F screened, G scr date,
    # H fr status, I fr int, J fr date, K partner status, L partner, M partner date,
    # N assess req, O sent, P returned, Q ret date, R ref, S offer init, T offer date,
    # U outcome, V notes
    ["Jordan Reyes", "Paralegal", date(2026, 5, 20), "Referral", "Referral",
     "Yes", date(2026, 5, 23), "Passed", "Amanda", date(2026, 5, 28),
     "Scheduled", "Marc Reich", None, "No", None, None, None, "Not Started",
     "No", None, None, "Strong paralegal; partner round scheduled."],
    ["Priya Nair", "Attorney – Data Breach", date(2026, 4, 15),
     "LinkedIn (Active)", "LinkedIn Active", "Yes", date(2026, 4, 18), None,
     None, None, None, None, None, "Yes", None, None, None, "Not Started",
     "No", None, None, "Screened; awaiting first-round scheduling."],
    ["Marcus Webb", "Partner-Level Associate Attorney", date(2026, 3, 1),
     "Internal", "Email", "Yes", date(2026, 3, 5), "Passed", "Tyler",
     date(2026, 3, 12), "Completed", "Ed Korsinsky", date(2026, 3, 20), "Yes",
     date(2026, 3, 22), "Yes", date(2026, 3, 28), "Completed", "Yes",
     date(2026, 4, 1), "Pending", "Offer extended; awaiting candidate response."],
    ["Sofia Klein", "Legal Assistant", date(2026, 2, 10), "University / School",
     "Email", "Yes", date(2026, 2, 12), "Passed", "Melissa", date(2026, 2, 18),
     "Completed", "Marc Reich", date(2026, 2, 25), "No", None, None, None,
     "Completed", "Yes", date(2026, 3, 2), "Accepted",
     "Hired – start date being finalized."],
]

for ri, row in enumerate(samples, start=2):
    for ci, val in enumerate(row, start=1):
        c = tr.cell(row=ri, column=ci, value=val)
        c.font = F_BASE
        c.border = BORDER
        c.alignment = left if ci in (1, 22) else center
        if ci in DATE_COLS and val is not None:
            c.number_format = "MM/DD/YYYY"

# ---- formulas + per-row formatting across full range ----------------------
for r in range(2, LAST + 1):
    # date number formats on all data rows
    for ci in DATE_COLS:
        cell = tr.cell(row=r, column=ci)
        cell.number_format = "MM/DD/YYYY"
        if cell.font is None or cell.font.name != "Calibri":
            cell.font = F_BASE
    # auto columns
    w = tr.cell(row=r, column=23, value=pipeline_formula(r))
    x = tr.cell(row=r, column=24, value=status_formula(r))
    y = tr.cell(row=r, column=25, value=days_formula(r))
    for cell in (w, x, y):
        cell.font = F_AUTO
        cell.alignment = center
    y.number_format = "0"

tr.freeze_panes = "C2"
tr.auto_filter.ref = f"A1:{get_column_letter(NCOL)}{LAST}"

# ---- data validation ------------------------------------------------------
dv_map = {
    "role": "B", "source": "D", "method": "E", "yesno": ["F", "N", "P", "S"],
    "status": ["H", "K"], "frint": "I", "partner": "L", "ref": "R",
    "outcome": "U",
}
for key, cols in dv_map.items():
    if isinstance(cols, str):
        cols = [cols]
    dv = DataValidation(type="list", formula1=VR[key], allow_blank=True,
                        showErrorMessage=True)
    dv.error = "Pick a value from the dropdown list."
    dv.errorTitle = "Invalid entry"
    dv.prompt = None
    for col in cols:
        dv.add(f"{col}2:{col}{LAST}")
    tr.add_data_validation(dv)

# ---- conditional formatting -----------------------------------------------
# NOTE: conditional-formatting fills must set BOTH colours (fg + bg); Excel
# uses bgColor for the visible fill in a differential format.
def cf_fill(hexv):
    return PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")

amber = cf_fill("FFEB9C")
red = cf_fill("FFC7CE")
FULL = f"A2:Y{LAST}"
cf_border = Border(left=Side(style="thin", color="BFBFBF"),
                   right=Side(style="thin", color="BFBFBF"),
                   top=Side(style="thin", color="BFBFBF"),
                   bottom=Side(style="thin", color="BFBFBF"))

# Rules are added highest-priority first (openpyxl assigns priority by order).
# 1) overdue: offer pending >14 days (red, flag Offer Initiated Date)
tr.conditional_formatting.add(
    f"T2:T{LAST}",
    FormulaRule(formula=['AND($U2="Pending",$T2<>"",TODAY()-$T2>14)'], fill=red))
# 2) overdue ambers
tr.conditional_formatting.add(
    f"F2:F{LAST}",
    FormulaRule(formula=['AND($A2<>"",$F2="",$C2<>"",TODAY()-$C2>10)'],
                fill=amber))
tr.conditional_formatting.add(
    f"J2:J{LAST}",
    FormulaRule(formula=['AND($F2="Yes",$G2<>"",$J2="",$X2="Active",'
                         '$H2<>"Rejected",$H2<>"Cancelled",TODAY()-$G2>7)'],
                fill=amber))
tr.conditional_formatting.add(
    f"Q2:Q{LAST}",
    FormulaRule(formula=['AND($O2<>"",$P2<>"Yes",$X2="Active",TODAY()-$O2>5)'],
                fill=amber))
# 3) offer outcome colour (cell)
tr.conditional_formatting.add(
    f"U2:U{LAST}", CellIsRule(operator="equal", formula=['"Accepted"'],
        fill=cf_fill("C6EFCE"), font=Font(color="006100", bold=True)))
tr.conditional_formatting.add(
    f"U2:U{LAST}", CellIsRule(operator="equal", formula=['"Declined"'],
        fill=cf_fill("FFC7CE"), font=Font(color="9C0006", bold=True)))
tr.conditional_formatting.add(
    f"U2:U{LAST}", CellIsRule(operator="equal", formula=['"Pending"'],
        fill=cf_fill("FFEB9C"), font=Font(color="9C6500", bold=True)))
# 4) active / closed text colour (font only, so row tint shows through)
tr.conditional_formatting.add(
    f"X2:X{LAST}", CellIsRule(operator="equal", formula=['"Closed"'],
        font=Font(color="808080", italic=True)))
tr.conditional_formatting.add(
    f"X2:X{LAST}", CellIsRule(operator="equal", formula=['"Active"'],
        font=Font(color="375623", bold=True)))
# 5) role colour coding — tint the whole candidate row by role
for role, hexv in ROLE_FILLS.items():
    tr.conditional_formatting.add(
        FULL, FormulaRule(formula=[f'$B2="{role}"'], fill=cf_fill(hexv)))
# 6) light borders on every row that has a candidate name (auto-expands)
tr.conditional_formatting.add(
    FULL, FormulaRule(formula=['$A2<>""'], border=cf_border))

# ==========================================================================
# DASHBOARD SHEET
# ==========================================================================
db = wb.create_sheet("Dashboard", 1)
db.sheet_view.showGridLines = False
for col, w in {"A": 30, "B": 12, "C": 14, "D": 14, "E": 16, "F": 4,
               "G": 22, "H": 16}.items():
    db.column_dimensions[col].width = w

NAMES = f"Tracker!$A$2:$A${LAST}"
ROLECOL = f"Tracker!$B$2:$B${LAST}"

def title(cell, text):
    db[cell] = text
    db[cell].font = F_TITLE

def subhead(rownum, text, span=3):
    c = db.cell(row=rownum, column=1, value=text)
    c.font = F_SUB
    c.fill = FILL_SUB
    c.alignment = left
    for k in range(2, span + 1):
        cc = db.cell(row=rownum, column=k)
        cc.fill = FILL_SUB

def th(rownum, labels, startcol=1):
    for i, lab in enumerate(labels):
        c = db.cell(row=rownum, column=startcol + i, value=lab)
        c.font = F_BOLD
        c.alignment = center if i else left
        c.border = BORDER
        c.fill = FILL_TOTAL

title("A1", "Recruiting Pipeline Dashboard")
db["A2"] = "Live summary – updates automatically as you edit the Tracker tab."
db["A2"].font = F_NOTE

# --- Section 1: Candidates by Role ---
subhead(4, "Candidates by Role")
th(5, ["Role / Position", "Count", "% of Total"])
r = 6
for i, role in enumerate(ROLES):
    lc = ls.cell(row=2 + i, column=1).coordinate  # Lists!A2..A7
    db.cell(row=r, column=1, value=f"=Lists!$A${2 + i}").font = F_BASE
    db.cell(row=r, column=2,
            value=f'=COUNTIF({ROLECOL},Lists!$A${2 + i})').font = F_BASE
    db.cell(row=r, column=3,
            value=f'=IFERROR(B{r}/$B${6 + len(ROLES)},0)').font = F_BASE
    db.cell(row=r, column=3).number_format = "0.0%"
    for cc in (1, 2, 3):
        db.cell(row=r, column=cc).border = BORDER
    db.cell(row=r, column=2).alignment = center
    db.cell(row=r, column=3).alignment = center
    r += 1
# total
db.cell(row=r, column=1, value="Total").font = F_BOLD
db.cell(row=r, column=2, value=f"=SUM(B6:B{r-1})").font = F_BOLD
db.cell(row=r, column=3, value=f"=IFERROR(B{r}/B{r},0)").font = F_BOLD
db.cell(row=r, column=3).number_format = "0.0%"
for cc in (1, 2, 3):
    db.cell(row=r, column=cc).border = BORDER
    db.cell(row=r, column=cc).fill = FILL_TOTAL
TOTAL_ROW = r

# --- Section 2: Pipeline Funnel & Conversion ---
fr0 = r + 3
subhead(fr0, "Pipeline Funnel & Conversion", span=4)
th(fr0 + 1, ["Stage", "Count", "% of Contacted", "Step Conversion"])
stages = [
    ("Contacted", f"=COUNTA({NAMES})"),
    ("Screened", f'=COUNTIF(Tracker!$F$2:$F${LAST},"Yes")'),
    ("First Round", f'=COUNTIF(Tracker!$H$2:$H${LAST},"<>")'),
    ("Partner Round", f'=COUNTIF(Tracker!$K$2:$K${LAST},"<>")'),
    ("Assessment", f'=COUNTIF(Tracker!$O$2:$O${LAST},"<>")'),
    ("Reference", f'=COUNTIF(Tracker!$R$2:$R${LAST},"In Progress")'
                  f'+COUNTIF(Tracker!$R$2:$R${LAST},"Completed")'),
    ("Offer", f'=COUNTIF(Tracker!$S$2:$S${LAST},"Yes")'),
    ("Hired (Accepted)", f'=COUNTIF(Tracker!$U$2:$U${LAST},"Accepted")'),
]
start = fr0 + 2
top = start  # Contacted row
for i, (label, formula) in enumerate(stages):
    rr = start + i
    db.cell(row=rr, column=1, value=label).font = F_BASE
    db.cell(row=rr, column=2, value=formula).font = F_BASE
    db.cell(row=rr, column=3,
            value=f"=IFERROR(B{rr}/$B${top},0)").font = F_BASE
    db.cell(row=rr, column=3).number_format = "0.0%"
    if i == 0:
        db.cell(row=rr, column=4, value="—").font = F_BASE
    else:
        db.cell(row=rr, column=4,
                value=f'=IFERROR(B{rr}/B{rr-1},"")').font = F_BASE
        db.cell(row=rr, column=4).number_format = "0.0%"
    for cc in (1, 2, 3, 4):
        db.cell(row=rr, column=cc).border = BORDER
        if cc > 1:
            db.cell(row=rr, column=cc).alignment = center
FUNNEL_END = start + len(stages) - 1

# --- Section 3: Status Summary (right side, col G/H) ---
subhead2_row = 4
sc = 7  # column G
db.cell(row=subhead2_row, column=sc, value="Status Summary").font = F_SUB
db.cell(row=subhead2_row, column=sc).fill = FILL_SUB
db.cell(row=subhead2_row, column=sc + 1).fill = FILL_SUB
status_rows = [
    ("Active Candidates", f'=COUNTIF(Tracker!$X$2:$X${LAST},"Active")'),
    ("Closed Candidates", f'=COUNTIF(Tracker!$X$2:$X${LAST},"Closed")'),
    ("Offers – Accepted", f'=COUNTIF(Tracker!$U$2:$U${LAST},"Accepted")'),
    ("Offers – Declined", f'=COUNTIF(Tracker!$U$2:$U${LAST},"Declined")'),
    ("Offers – Pending", f'=COUNTIF(Tracker!$U$2:$U${LAST},"Pending")'),
]
for i, (label, formula) in enumerate(status_rows):
    rr = subhead2_row + 1 + i
    db.cell(row=rr, column=sc, value=label).font = F_BASE
    db.cell(row=rr, column=sc + 1, value=formula).font = F_BASE
    db.cell(row=rr, column=sc + 1).alignment = center
    db.cell(row=rr, column=sc).border = BORDER
    db.cell(row=rr, column=sc + 1).border = BORDER

# --- Time in Pipeline (right side, below status) ---
tp_row = subhead2_row + len(status_rows) + 3
db.cell(row=tp_row, column=sc, value="Time in Pipeline").font = F_SUB
db.cell(row=tp_row, column=sc).fill = FILL_SUB
db.cell(row=tp_row, column=sc + 1).fill = FILL_SUB
# MAX(IF(...)) array formula instead of MAXIFS (works on all Excel versions)
time_rows = [
    ("Avg Days (Active)",
     f'=IFERROR(ROUND(AVERAGEIFS(Tracker!$Y$2:$Y${LAST},'
     f'Tracker!$X$2:$X${LAST},"Active"),0),0)', False),
    ("Longest Active (Days)",
     f'=IFERROR(MAX(IF(Tracker!$X$2:$X${LAST}="Active",'
     f'Tracker!$Y$2:$Y${LAST})),0)', True),
    ("Total Candidates", f"=COUNTA({NAMES})", False),
]
for i, (label, formula, is_array) in enumerate(time_rows):
    rr = tp_row + 1 + i
    db.cell(row=rr, column=sc, value=label).font = F_BASE
    fcell = db.cell(row=rr, column=sc + 1)
    if is_array:
        fcell.value = ArrayFormula(fcell.coordinate, formula)
    else:
        fcell.value = formula
    fcell.font = F_BASE
    db.cell(row=rr, column=sc + 1).alignment = center
    db.cell(row=rr, column=sc).border = BORDER
    db.cell(row=rr, column=sc + 1).border = BORDER
db.cell(row=tp_row + 1 + len(time_rows), column=sc,
        value='"Days in Pipeline" = days since Date Contacted.').font = F_NOTE

# --- Legend ---
leg = FUNNEL_END + 3
db.cell(row=leg, column=1, value="Legend – Conditional Formatting").font = F_SUB
db.cell(row=leg, column=1).fill = FILL_SUB
for k in (2, 3):
    db.cell(row=leg, column=k).fill = FILL_SUB
legend = [
    ("FFEB9C", "Amber = stage overdue (needs follow-up)"),
    ("FFC7CE", "Red = offer pending too long (>14 days)"),
    ("C6EFCE", "Green = offer accepted"),
    ("D9D9D9", "Grey = candidate closed"),
]
for i, (hexv, txt) in enumerate(legend):
    rr = leg + 1 + i
    db.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=hexv)
    db.cell(row=rr, column=1).border = BORDER
    db.cell(row=rr, column=2, value=txt).font = F_BASE
roles_leg = leg + 1 + len(legend) + 1
db.cell(row=roles_leg, column=1, value="Role colours:").font = F_BOLD
for i, (role, hexv) in enumerate(ROLE_FILLS.items()):
    rr = roles_leg + 1 + i
    db.cell(row=rr, column=1).fill = PatternFill("solid", fgColor=hexv)
    db.cell(row=rr, column=1).border = BORDER
    db.cell(row=rr, column=2, value=role).font = F_BASE

# ==========================================================================
# READ ME SHEET
# ==========================================================================
rm = wb.create_sheet("Read Me", 0)
rm.sheet_view.showGridLines = False
rm.column_dimensions["A"].width = 2
rm.column_dimensions["B"].width = 30
rm.column_dimensions["C"].width = 92
_r = {"n": 1}

def put(text, font, fill=None, height=None, indent=1, valign="center"):
    c = rm.cell(row=_r["n"], column=2, value=text)
    c.font = font
    if fill:
        c.fill = fill
    c.alignment = Alignment(vertical=valign, wrap_text=True, indent=indent)
    rm.merge_cells(start_row=_r["n"], start_column=2, end_row=_r["n"], end_column=3)
    if height:
        rm.row_dimensions[_r["n"]].height = height
    _r["n"] += 1

def section(text):
    c = rm.cell(row=_r["n"], column=2, value=text)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = FILL_SUB
    c.alignment = Alignment(vertical="center", indent=1)
    rm.cell(row=_r["n"], column=3).fill = FILL_SUB
    rm.merge_cells(start_row=_r["n"], start_column=2, end_row=_r["n"], end_column=3)
    rm.row_dimensions[_r["n"]].height = 22
    _r["n"] += 1

def bullet(text):
    put("•  " + text, Font(name="Calibri", size=11), valign="top")

def gap():
    _r["n"] += 1

def swatch(label, hexv, desc, fontcolor="000000"):
    b = rm.cell(row=_r["n"], column=2, value=label)
    b.fill = PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")
    b.font = Font(name="Calibri", size=11, bold=True, color=fontcolor)
    b.alignment = Alignment(vertical="center", horizontal="center")
    b.border = BORDER
    c = rm.cell(row=_r["n"], column=3, value=desc)
    c.font = Font(name="Calibri", size=11)
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    _r["n"] += 1

put("Recruiting Tracker — Read Me",
    Font(name="Calibri", size=18, bold=True, color="1F4E78"), height=28)
put("How the workbook is organised and how to use it day-to-day.",
    Font(name="Calibri", size=11, italic=True, color="808080"))
gap()

section("1.  The Tabs")
bullet("Tracker — one row per candidate. Type directly into the cells; the green "
       "columns on the far right fill in automatically.")
bullet("Dashboard — a live summary (candidates by role, pipeline funnel with "
       "conversion rates, status, and time in pipeline). It updates by itself as "
       "you edit the Tracker.")
bullet("Lists — the menu of choices behind every dropdown. Edit values here to "
       "change the available options.")
bullet("Read Me — this tab.")
gap()

section("2.  Adding a Candidate")
bullet("On the Tracker tab, type the candidate's name in the next empty row of "
       "column A.")
bullet("Fill the categorical fields with the dropdown arrow that appears when you "
       "click the cell (Role, Source, Screened?, statuses, interviewers, etc.).")
bullet("Enter dates as MM/DD/YYYY.")
bullet("Leave the last three columns alone — Pipeline Stage, Active / Closed and "
       "Days in Pipeline calculate themselves.")
gap()

section("3.  The Automatic Columns (green headers)")
bullet("Pipeline Stage — the furthest stage reached: Contacted → Screened → First "
       "Round → Partner → Assessment → Reference → Offer → Hired.")
bullet("Active / Closed — shows \"Closed\" once an offer is Accepted or Declined, "
       "or a candidate is Rejected; otherwise \"Active\".")
bullet("Days in Pipeline — number of days since Date Contacted.")
gap()

section("4.  Colour Coding")
bullet("Each role tints its whole row a distinct colour:")
for role, hexv in ROLE_FILLS.items():
    swatch("", hexv, role)
gap()
bullet("Offer Outcome cell colour: green = Accepted, red = Declined, amber = Pending.")
bullet("Attention flags appear automatically when a stage is overdue:")
swatch("Overdue", "FFEB9C", "Amber — a stage is taking too long and needs a nudge.")
swatch("Too long", "FFC7CE", "Red — an offer has been pending more than 14 days.",
       fontcolor="9C0006")
gap()

section("5.  When the Attention Flags Trigger")
bullet("Contacted but still not screened after 10 days (amber on Screened?).")
bullet("Screened but no first-round date after 7 days (amber on First Round Date).")
bullet("Assessment sent but not returned after 5 days (amber on Return Date).")
bullet("Offer still Pending more than 14 days (red on Offer Initiated Date).")
bullet("These day thresholds are easy to change on request.")
gap()

section("6.  Filtering & Sorting")
bullet("Every column header has a filter arrow — filter by Role, Pipeline Stage, "
       "Interviewer, Active / Closed, and more.")
bullet("To see only open candidates, filter Active / Closed = Active.")
bullet("The header row and the Name + Role columns stay frozen as you scroll.")
gap()

section("7.  Changing Dropdown Options")
bullet("Open the Lists tab and edit the values under the matching heading "
       "(for example, add a new interviewer).")
bullet("If you add a value beyond the current coloured range, the dropdown won't "
       "pick it up automatically — ask to have the range widened.")
gap()

section("8.  Before You Start")
bullet("Rows 2–5 of the Tracker are sample candidates included to show the colours "
       "and flags in action. Delete them before entering real data.")

wb.active = wb.sheetnames.index("Tracker")

# ==========================================================================
wb.save(OUT)
print("Saved", OUT)
print("Total row:", TOTAL_ROW, "Funnel:", start, "-", FUNNEL_END)
