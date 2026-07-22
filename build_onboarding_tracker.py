# -*- coding: utf-8 -*-
"""Build the New Hire Onboarding Tracker workbook: Onboarding, Lists, Read Me.

The data area is a real Excel table (OnboardingTable) so Power Automate's
"List rows present in a table" action can read it for the Teams/Outlook
calendar integration (see Onboarding_Calendar_Setup.md).
"""
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\DSamson\.claude\Settlement\Onboarding_Tracker.xlsx"
TABLE_LAST = 151   # table covers rows 1-151 (150 hire rows) for Power Automate
LAST = 400         # dropdowns / date formats extend here so the table can grow

# ---- dropdown value lists (options provided by HR, Jul 2026) --------------
POSITIONS = ["Paralegal", "Legal Assistant", "Jr. Attorney",
             "Mid-Senior Level Attorney", "Other"]
LOCATIONS = ["On-Site", "Remote"]
YESNO = ["Yes", "No"]
MANAGERS = ["Amanda", "Mark", "Ed", "Sam", "Jillian", "Michael", "Keith"]

# ---- fonts / fills --------------------------------------------------------
F_BASE = Font(name="Calibri", size=11)
F_BOLD = Font(name="Calibri", size=11, bold=True)
F_HEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_AUTO = Font(name="Calibri", size=11, italic=True, color="FFFFFF")
F_NOTE = Font(name="Calibri", size=9, italic=True, color="808080")

FILL_HEAD = PatternFill("solid", fgColor="1F4E78")
FILL_AUTOHEAD = PatternFill("solid", fgColor="375623")
FILL_SUB = PatternFill("solid", fgColor="2E75B6")
FILL_LISTHEAD = PatternFill("solid", fgColor="808080")
FILL_INPUT = PatternFill("solid", fgColor="FFF2CC")

POSITION_FILLS = {
    "Paralegal": "BDD7EE",
    "Legal Assistant": "C6E0B4",
    "Jr. Attorney": "FFE699",
    "Mid-Senior Level Attorney": "B4A7D6",
    "Other": "D9D9D9",
}

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ==========================================================================
# LISTS SHEET
# ==========================================================================
ls = wb.active
ls.title = "Lists"
list_cols = [
    ("Position", POSITIONS),
    ("Work Location", LOCATIONS),
    ("Laptop Needed", YESNO),
    ("Reporting To", MANAGERS),
]
for ci, (hdr, vals) in enumerate(list_cols, start=1):
    c = ls.cell(row=1, column=ci, value=hdr)
    c.font = F_HEAD
    c.fill = FILL_LISTHEAD
    c.alignment = center
    ls.column_dimensions[get_column_letter(ci)].width = max(18, len(hdr) + 2)
    for ri, v in enumerate(vals, start=2):
        ls.cell(row=ri, column=ci, value=v).font = F_BASE

# optional manager-email column (used by the calendar flow to invite managers)
c = ls.cell(row=1, column=5, value="Manager Email (optional)")
c.font = F_HEAD
c.fill = FILL_LISTHEAD
c.alignment = center
ls.column_dimensions["E"].width = 34
for ri in range(2, 2 + len(MANAGERS)):
    ec = ls.cell(row=ri, column=5)
    ec.fill = FILL_INPUT
    ec.font = F_BASE

note_row = 2 + len(MANAGERS) + 2
ls.cell(row=note_row, column=1,
        value="Reference lists that drive the dropdowns on the Onboarding tab. "
              "Edit a value here and every dropdown updates. Position and "
              "manager options provided by HR (Jul 2026).").font = F_NOTE
ls.cell(row=note_row + 1, column=1,
        value="Manager Email (yellow cells): optional - fill in if you want "
              "the Power Automate calendar flow to invite each manager to "
              "their new hire's event (see Onboarding_Calendar_Setup.md)."
        ).font = F_NOTE

def rng(col_idx, n):
    L = get_column_letter(col_idx)
    return f"Lists!${L}$2:${L}${1 + n}"

VR = {
    "position": rng(1, len(POSITIONS)),
    "location": rng(2, len(LOCATIONS)),
    "yesno": rng(3, len(YESNO)),
    "manager": rng(4, len(MANAGERS)),
}

# ==========================================================================
# ONBOARDING SHEET
# ==========================================================================
onb = wb.create_sheet("Onboarding", 0)

HEADERS = [
    ("Candidate Name", 24, None),        # A
    ("Position", 27, "position"),        # B
    ("Start Date", 14, "date"),          # C
    ("Work Location", 14, "location"),   # D
    ("Laptop Needed", 14, "yesno"),      # E
    ("Reporting To", 14, "manager"),     # F
    ("Email", 32, None),                 # G
    ("Added to Calendar", 16, "auto"),   # H - written by the Power Automate flow
]
NCOL = len(HEADERS)

onb.row_dimensions[1].height = 32
for ci, (name, width, kind) in enumerate(HEADERS, start=1):
    c = onb.cell(row=1, column=ci, value=name)
    c.font = F_AUTO if kind == "auto" else F_HEAD
    c.fill = FILL_AUTOHEAD if kind == "auto" else FILL_HEAD
    c.alignment = center
    c.border = BORDER
    onb.column_dimensions[get_column_letter(ci)].width = width

# ---- sample row (delete before real use - noted on Read Me) ---------------
sample = ["Dana Whitfield", "Paralegal", date(2026, 8, 3), "On-Site", "Yes",
          "Amanda", "dana.whitfield@gmail.com", "Yes"]
for ci, val in enumerate(sample, start=1):
    onb.cell(row=2, column=ci, value=val)

# ---- per-row formatting across the full entry range -----------------------
for r in range(2, LAST + 1):
    for ci in range(1, NCOL + 1):
        cell = onb.cell(row=r, column=ci)
        cell.font = F_BASE
        cell.alignment = left if ci in (1, 7) else center
        if ci == 3:
            cell.number_format = "MM/DD/YYYY"

onb.freeze_panes = "B2"

# ---- Excel table (required by Power Automate's Excel connector) -----------
tab = Table(displayName="OnboardingTable", ref=f"A1:H{TABLE_LAST}")
tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True,
                                    showColumnStripes=False, showFirstColumn=False,
                                    showLastColumn=False)
onb.add_table(tab)

# ---- data validation ------------------------------------------------------
for key, col in [("position", "B"), ("location", "D"), ("yesno", "E"),
                 ("manager", "F"), ("yesno", "H")]:
    dv = DataValidation(type="list", formula1=VR[key], allow_blank=True,
                        showErrorMessage=True)
    dv.error = "Pick a value from the dropdown list."
    dv.errorTitle = "Invalid entry"
    dv.add(f"{col}2:{col}{LAST}")
    onb.add_data_validation(dv)

dv_date = DataValidation(type="date", operator="greaterThan",
                         formula1="DATE(2020,1,1)", allow_blank=True,
                         showErrorMessage=True, showInputMessage=True)
dv_date.error = "Enter a real date (MM/DD/YYYY)."
dv_date.errorTitle = "Invalid date"
dv_date.promptTitle = "Start Date"
dv_date.prompt = "First day of work - MM/DD/YYYY."
dv_date.add(f"C2:C{LAST}")
onb.add_data_validation(dv_date)

# ---- conditional formatting -----------------------------------------------
# NOTE: conditional-formatting fills must set BOTH colours (fg + bg); Excel
# uses bgColor for the visible fill in a differential format.
def cf_fill(hexv):
    return PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")

FULL = f"A2:H{LAST}"
cf_border = Border(left=thin, right=thin, top=thin, bottom=thin)

# 1) start date within the next 7 days -> amber (prep laptop / desk / accounts)
onb.conditional_formatting.add(
    f"C2:C{LAST}",
    FormulaRule(formula=['AND($A2<>"",$C2<>"",$C2>=TODAY(),$C2-TODAY()<=7)'],
                fill=cf_fill("FFEB9C"), font=Font(color="9C6500", bold=True)))
# 2) already started -> grey italic date
onb.conditional_formatting.add(
    f"C2:C{LAST}",
    FormulaRule(formula=['AND($C2<>"",$C2<TODAY())'],
                font=Font(color="808080", italic=True)))
# 3) laptop needed -> orange flag for IT
onb.conditional_formatting.add(
    f"E2:E{LAST}", CellIsRule(operator="equal", formula=['"Yes"'],
        fill=cf_fill("FCE4D6"), font=Font(color="C55A11", bold=True)))
# 4) on the calendar -> green
onb.conditional_formatting.add(
    f"H2:H{LAST}", CellIsRule(operator="equal", formula=['"Yes"'],
        fill=cf_fill("C6EFCE"), font=Font(color="006100", bold=True)))
# 5) position colour coding - tint the whole row by position
for pos, hexv in POSITION_FILLS.items():
    onb.conditional_formatting.add(
        FULL, FormulaRule(formula=[f'$B2="{pos}"'], fill=cf_fill(hexv)))
# 6) light borders on every row that has a candidate name (auto-expands)
onb.conditional_formatting.add(
    FULL, FormulaRule(formula=['$A2<>""'], border=cf_border))

# ==========================================================================
# READ ME SHEET
# ==========================================================================
rm = wb.create_sheet("Read Me", 0)
rm.sheet_view.showGridLines = False
rm.column_dimensions["A"].width = 2
rm.column_dimensions["B"].width = 30
rm.column_dimensions["C"].width = 92
_r = {"n": 1}

def put(text, font, fill=None, height=None, indent=1, valign="center"):
    c = rm.cell(row=_r["n"], column=2, value=text)
    c.font = font
    if fill:
        c.fill = fill
    c.alignment = Alignment(vertical=valign, wrap_text=True, indent=indent)
    rm.merge_cells(start_row=_r["n"], start_column=2, end_row=_r["n"], end_column=3)
    if height:
        rm.row_dimensions[_r["n"]].height = height
    _r["n"] += 1

def section(text):
    c = rm.cell(row=_r["n"], column=2, value=text)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = FILL_SUB
    c.alignment = Alignment(vertical="center", indent=1)
    rm.cell(row=_r["n"], column=3).fill = FILL_SUB
    rm.merge_cells(start_row=_r["n"], start_column=2, end_row=_r["n"], end_column=3)
    rm.row_dimensions[_r["n"]].height = 22
    _r["n"] += 1

def bullet(text):
    put("-  " + text, Font(name="Calibri", size=11), valign="top")

def gap():
    _r["n"] += 1

def swatch(label, hexv, desc, fontcolor="000000"):
    b = rm.cell(row=_r["n"], column=2, value=label)
    b.fill = PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")
    b.font = Font(name="Calibri", size=11, bold=True, color=fontcolor)
    b.alignment = Alignment(vertical="center", horizontal="center")
    b.border = BORDER
    c = rm.cell(row=_r["n"], column=3, value=desc)
    c.font = Font(name="Calibri", size=11)
    c.alignment = Alignment(vertical="center", wrap_text=True, indent=1)
    _r["n"] += 1

put("New Hire Onboarding Tracker - Read Me",
    Font(name="Calibri", size=18, bold=True, color="1F4E78"), height=28)
put("One row per new hire. How the workbook is organised and how to use it.",
    Font(name="Calibri", size=11, italic=True, color="808080"))
gap()

section("1.  The Tabs")
bullet("Onboarding - one row per new hire. You type columns A-G; the green "
       "column H (Added to Calendar) fills in automatically once the "
       "calendar flow is connected.")
bullet("Lists - the menu of choices behind every dropdown. Edit values here "
       "to change the available options.")
bullet("Read Me - this tab.")
gap()

section("2.  Adding a New Hire")
bullet("Type the candidate's name in the next empty row of column A (inside "
       "the striped table area - the table grows with you).")
bullet("Position, Work Location, Laptop Needed and Reporting To are dropdowns "
       "- click the cell and use the arrow that appears.")
bullet("Enter Start Date as MM/DD/YYYY and type the personal email in the "
       "Email column.")
bullet("Row 2 (Dana Whitfield) is a sample showing the expected format - "
       "delete it before entering real hires.")
gap()

section("3.  The 'Added to Calendar' Column (green header)")
bullet("Leave it blank when you add a hire. The Power Automate flow creates "
       "the Outlook / Teams calendar event for the start date, then writes "
       "'Yes' here so the event is never duplicated.")
bullet("To keep a hire OFF the calendar, set it to 'Yes' yourself before the "
       "flow's next run. Clearing the cell makes the flow create the event "
       "on its next pass.")
bullet("Setup instructions: Onboarding_Calendar_Setup.md (kept alongside "
       "this workbook).")
gap()

section("4.  Colour Coding")
bullet("Each position tints its whole row:")
for pos, hexv in POSITION_FILLS.items():
    swatch("", hexv, pos)
gap()
swatch("Soon", "FFEB9C", "Amber Start Date - starts within the next 7 days "
       "(prep laptop, desk, accounts).", fontcolor="9C6500")
swatch("Laptop", "FCE4D6", "Orange 'Yes' in Laptop Needed - IT needs to "
       "provision a machine.", fontcolor="C55A11")
swatch("Booked", "C6EFCE", "Green 'Yes' in Added to Calendar - the start "
       "date is on the calendar.", fontcolor="006100")
bullet("A grey italic Start Date means the person has already started.")
gap()

section("5.  Changing Dropdown Options")
bullet("Open the Lists tab and edit the values under the matching heading "
       "(for example, swap a manager's name).")
bullet("If you need to ADD an option beyond the current list (an 8th "
       "manager, a new position), type it on Lists and extend the range via "
       "Data > Data Validation - or ask to have it widened.")
bullet("The yellow Manager Email column on Lists is optional - fill it in "
       "if the calendar flow should invite managers to their hires' events.")
gap()

section("6.  The Table (for the calendar integration)")
bullet("The data area is an Excel table named 'OnboardingTable'. Power "
       "Automate reads rows from this table - do not convert it to a normal "
       "range or rename it.")
bullet("Keep this file in the Teams channel's Files tab (SharePoint) once "
       "the flow is connected; moving or renaming it breaks the flow.")

wb.active = wb.sheetnames.index("Onboarding")
wb.save(OUT)
print("Saved", OUT)
