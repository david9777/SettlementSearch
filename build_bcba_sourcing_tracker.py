# -*- coding: utf-8 -*-
"""Build the BCBA Sourcing Tracker workbook (NC + MD).

A Delaware-style recruiting/sourcing roster built ONLY on public professional
licensing data. Intentionally EXCLUDES home address, age, and personal
phone/email — see the Read Me tab for why.
"""
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\DSamson\.claude\Settlement\BCBA_Sourcing_Tracker.xlsx"
LAST = 600  # last row that gets dropdowns / formulas / formatting

# ---- dropdown value lists -------------------------------------------------
CREDENTIALS = ["BCBA-D", "BCBA", "BCaBA", "RBT",
               "LBA (state license)", "LABA (state assistant)"]
STATES = ["NC", "MD"]
STATUS = ["Active", "Expired", "Inactive", "Suspended", "Unknown"]
SOURCE = ["BACB Registry", "NC LBA Board", "MD Board (BOPC)", "Multiple"]
OUTREACH = ["Not Contacted", "Contacted", "Responded", "Screening",
            "In Pipeline", "Not Interested", "Do Not Contact"]

# ---- fonts / fills --------------------------------------------------------
F_BASE = Font(name="Calibri", size=11)
F_BOLD = Font(name="Calibri", size=11, bold=True)
F_HEAD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
F_AUTO = Font(name="Calibri", size=11, italic=True, color="3F3F3F")
F_TITLE = Font(name="Calibri", size=16, bold=True, color="1F4E78")
F_SUB = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
F_NOTE = Font(name="Calibri", size=9, italic=True, color="808080")

FILL_HEAD = PatternFill("solid", fgColor="1F4E78")
FILL_AUTOHEAD = PatternFill("solid", fgColor="375623")
FILL_SUB = PatternFill("solid", fgColor="2E75B6")
FILL_TOTAL = PatternFill("solid", fgColor="DDEBF7")
FILL_LISTHEAD = PatternFill("solid", fgColor="808080")

STATE_FILLS = {"NC": "BDD7EE", "MD": "C6E0B4"}

thin = Side(style="thin", color="BFBFBF")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

wb = Workbook()

# ==========================================================================
# LISTS SHEET
# ==========================================================================
ls = wb.active
ls.title = "Lists"
list_cols = [
    ("Credential", CREDENTIALS),
    ("State", STATES),
    ("License Status", STATUS),
    ("Source Verified", SOURCE),
    ("Outreach Status", OUTREACH),
]
for ci, (hdr, vals) in enumerate(list_cols, start=1):
    c = ls.cell(row=1, column=ci, value=hdr)
    c.font = F_HEAD
    c.fill = FILL_LISTHEAD
    c.alignment = center
    ls.column_dimensions[get_column_letter(ci)].width = max(18, len(hdr) + 2)
    for ri, v in enumerate(vals, start=2):
        ls.cell(row=ri, column=ci, value=v).font = F_BASE
ls.cell(row=len(max(list_cols, key=lambda t: len(t[1]))[1]) + 3, column=1,
        value="Reference lists that drive the dropdowns on the Candidates tab. "
              "Edit values here to change the available options.").font = F_NOTE


def rng(col_idx, n):
    L = get_column_letter(col_idx)
    return f"Lists!${L}$2:${L}${1 + n}"


VR = {
    "cred": rng(1, len(CREDENTIALS)),
    "state": rng(2, len(STATES)),
    "status": rng(3, len(STATUS)),
    "source": rng(4, len(SOURCE)),
    "outreach": rng(5, len(OUTREACH)),
}

# ==========================================================================
# CANDIDATES SHEET
# ==========================================================================
tr = wb.create_sheet("Candidates", 0)

HEADERS = [
    ("Name", 24, None),
    ("Credential", 20, "cred"),
    ("State", 8, "state"),
    ("License / Cert #", 16, None),
    ("License / Cert Issue Date", 16, "date"),
    ("Years Licensed", 12, "auto"),
    ("License Status", 14, "status"),
    ("City", 18, None),
    ("Public Employer / Clinic", 26, None),
    ("LinkedIn URL", 30, None),
    ("Professional Contact (public office/work only)", 30, None),
    ("Source Verified", 16, "source"),
    ("Outreach Status", 16, "outreach"),
    ("Last Contact Date", 15, "date"),
    ("Notes", 34, None),
]
NCOL = len(HEADERS)

tr.row_dimensions[1].height = 42
for ci, (name, width, kind) in enumerate(HEADERS, start=1):
    c = tr.cell(row=1, column=ci, value=name)
    c.font = F_AUTO if kind == "auto" else F_HEAD
    c.fill = FILL_AUTOHEAD if kind == "auto" else FILL_HEAD
    c.alignment = center
    c.border = BORDER
    tr.column_dimensions[get_column_letter(ci)].width = width

DATE_COLS = [ci for ci, (_, _, k) in enumerate(HEADERS, start=1) if k == "date"]


def years_formula(r):
    # integer years since license issue date; blank until a date is entered
    return f'=IF($E{r}="","",DATEDIF($E{r},TODAY(),"Y"))'


# ---- illustrative sample rows (DELETE before real use) --------------------
samples = [
    ["Sample Candidate — delete", "BCBA", "NC", "1-00-0000",
     date(2018, 6, 1), None, "Active", "Raleigh", "Example ABA Group",
     "", "", "BACB Registry", "Not Contacted", None,
     "Illustrative row showing how to fill the sheet — delete before use."],
    ["Sample Candidate — delete", "LBA (state license)", "MD", "LBA-0000",
     date(2021, 9, 15), None, "Active", "Baltimore", "Example Behavioral Health",
     "", "", "MD Board (BOPC)", "Not Contacted", None,
     "Illustrative row showing how to fill the sheet — delete before use."],
]
for ri, row in enumerate(samples, start=2):
    for ci, val in enumerate(row, start=1):
        c = tr.cell(row=ri, column=ci, value=val)
        c.font = F_BASE
        c.border = BORDER
        c.alignment = left if ci in (1, 9, 10, 11, 15) else center
        if ci in DATE_COLS and val is not None:
            c.number_format = "MM/DD/YYYY"

# ---- formulas + per-row formatting across full range ----------------------
for r in range(2, LAST + 1):
    for ci in DATE_COLS:
        cell = tr.cell(row=r, column=ci)
        cell.number_format = "MM/DD/YYYY"
    f = tr.cell(row=r, column=6, value=years_formula(r))
    f.font = F_AUTO
    f.alignment = center
    f.number_format = "0"

tr.freeze_panes = "B2"
tr.auto_filter.ref = f"A1:{get_column_letter(NCOL)}{LAST}"

# ---- data validation ------------------------------------------------------
dv_map = {
    "cred": "B", "state": "C", "status": "G", "source": "L", "outreach": "M",
}
for key, col in dv_map.items():
    dv = DataValidation(type="list", formula1=VR[key], allow_blank=True,
                        showErrorMessage=True)
    dv.error = "Pick a value from the dropdown list."
    dv.errorTitle = "Invalid entry"
    dv.add(f"{col}2:{col}{LAST}")
    tr.add_data_validation(dv)


# ---- conditional formatting -----------------------------------------------
def cf_fill(hexv):
    return PatternFill(start_color=hexv, end_color=hexv, fill_type="solid")


FULL = f"A2:{get_column_letter(NCOL)}{LAST}"

# state row tint
for st, hexv in STATE_FILLS.items():
    tr.conditional_formatting.add(
        FULL, FormulaRule(formula=[f'$C2="{st}"'], fill=cf_fill(hexv)))

# outreach status cell colour
tr.conditional_formatting.add(
    f"M2:M{LAST}", CellIsRule(operator="equal", formula=['"In Pipeline"'],
        fill=cf_fill("C6EFCE"), font=Font(color="006100", bold=True)))
tr.conditional_formatting.add(
    f"M2:M{LAST}", CellIsRule(operator="equal", formula=['"Responded"'],
        fill=cf_fill("FFEB9C"), font=Font(color="9C6500", bold=True)))
tr.conditional_formatting.add(
    f"M2:M{LAST}", CellIsRule(operator="equal", formula=['"Do Not Contact"'],
        fill=cf_fill("FFC7CE"), font=Font(color="9C0006", bold=True)))
tr.conditional_formatting.add(
    f"M2:M{LAST}", CellIsRule(operator="equal", formula=['"Not Interested"'],
        font=Font(color="808080", italic=True)))

# light borders on every named row
cf_border = Border(left=Side(style="thin", color="BFBFBF"),
                   right=Side(style="thin", color="BFBFBF"),
                   top=Side(style="thin", color="BFBFBF"),
                   bottom=Side(style="thin", color="BFBFBF"))
tr.conditional_formatting.add(
    FULL, FormulaRule(formula=['$A2<>""'], border=cf_border))

# ==========================================================================
# DASHBOARD SHEET
# ==========================================================================
db = wb.create_sheet("Dashboard", 1)
db.sheet_view.showGridLines = False
for col, w in {"A": 30, "B": 12, "C": 14, "D": 4, "E": 26, "F": 12}.items():
    db.column_dimensions[col].width = w

STCOL = f"Candidates!$C$2:$C${LAST}"
CREDCOL = f"Candidates!$B$2:$B${LAST}"
STATUSCOL = f"Candidates!$G$2:$G${LAST}"
OUTCOL = f"Candidates!$M$2:$M${LAST}"
YRCOL = f"Candidates!$F$2:$F${LAST}"
NAMES = f"Candidates!$A$2:$A${LAST}"


def subhead(rownum, text, col=1, span=3):
    c = db.cell(row=rownum, column=col, value=text)
    c.font = F_SUB
    c.fill = FILL_SUB
    c.alignment = left
    for k in range(col + 1, col + span):
        db.cell(row=rownum, column=k).fill = FILL_SUB


def th(rownum, labels, startcol=1):
    for i, lab in enumerate(labels):
        c = db.cell(row=rownum, column=startcol + i, value=lab)
        c.font = F_BOLD
        c.alignment = center if i else left
        c.border = BORDER
        c.fill = FILL_TOTAL


db["A1"] = "BCBA Sourcing Dashboard — NC + MD"
db["A1"].font = F_TITLE
db["A2"] = "Live summary — updates automatically as you edit the Candidates tab."
db["A2"].font = F_NOTE


def simple_table(start_row, title_text, pairs, count_range=None, is_year=False):
    subhead(start_row, title_text)
    th(start_row + 1, [title_text.split(" by ")[-1] if " by " in title_text
                       else "Category", "Count"])
    r = start_row + 2
    for label, crit in pairs:
        db.cell(row=r, column=1, value=label).font = F_BASE
        if is_year:
            db.cell(row=r, column=2, value=crit).font = F_BASE
        else:
            db.cell(row=r, column=2,
                    value=f'=COUNTIF({count_range},"{crit}")').font = F_BASE
        db.cell(row=r, column=1).border = BORDER
        db.cell(row=r, column=2).border = BORDER
        db.cell(row=r, column=2).alignment = center
        r += 1
    db.cell(row=r, column=1, value="Total").font = F_BOLD
    db.cell(row=r, column=2, value=f"=SUM(B{start_row + 2}:B{r - 1})").font = F_BOLD
    for cc in (1, 2):
        db.cell(row=r, column=cc).border = BORDER
        db.cell(row=r, column=cc).fill = FILL_TOTAL
    return r + 3


row = 4
row = simple_table(row, "Candidates by State",
                   [(s, s) for s in STATES], STCOL)
row = simple_table(row, "Candidates by Credential",
                   [(c, c) for c in CREDENTIALS], CREDCOL)
row = simple_table(row, "Candidates by License Status",
                   [(s, s) for s in STATUS], STATUSCOL)

# right-side column: outreach + years-licensed buckets
subhead(4, "Outreach Status", col=5, span=2)
th(5, ["Status", "Count"], startcol=5)
rr = 6
for s in OUTREACH:
    db.cell(row=rr, column=5, value=s).font = F_BASE
    db.cell(row=rr, column=6, value=f'=COUNTIF({OUTCOL},"{s}")').font = F_BASE
    db.cell(row=rr, column=5).border = BORDER
    db.cell(row=rr, column=6).border = BORDER
    db.cell(row=rr, column=6).alignment = center
    rr += 1

yr0 = rr + 2
subhead(yr0, "Years Licensed", col=5, span=2)
th(yr0 + 1, ["Range", "Count"], startcol=5)
buckets = [
    ("0–2 years", f'=COUNTIFS({YRCOL},">=0",{YRCOL},"<=2")'),
    ("3–5 years", f'=COUNTIFS({YRCOL},">=3",{YRCOL},"<=5")'),
    ("6–10 years", f'=COUNTIFS({YRCOL},">=6",{YRCOL},"<=10")'),
    ("11+ years", f'=COUNTIF({YRCOL},">=11")'),
]
rr = yr0 + 2
for label, formula in buckets:
    db.cell(row=rr, column=5, value=label).font = F_BASE
    db.cell(row=rr, column=6, value=formula).font = F_BASE
    db.cell(row=rr, column=5).border = BORDER
    db.cell(row=rr, column=6).border = BORDER
    db.cell(row=rr, column=6).alignment = center
    rr += 1

# ==========================================================================
# READ ME SHEET
# ==========================================================================
rm = wb.create_sheet("Read Me", 0)
rm.sheet_view.showGridLines = False
rm.column_dimensions["A"].width = 2
rm.column_dimensions["B"].width = 30
rm.column_dimensions["C"].width = 92
_r = {"n": 1}


def put(text, font, fill=None, height=None, indent=1, valign="center"):
    c = rm.cell(row=_r["n"], column=2, value=text)
    c.font = font
    if fill:
        c.fill = fill
    c.alignment = Alignment(vertical=valign, wrap_text=True, indent=indent)
    rm.merge_cells(start_row=_r["n"], start_column=2, end_row=_r["n"], end_column=3)
    if height:
        rm.row_dimensions[_r["n"]].height = height
    _r["n"] += 1


def section(text):
    c = rm.cell(row=_r["n"], column=2, value=text)
    c.font = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    c.fill = FILL_SUB
    c.alignment = Alignment(vertical="center", indent=1)
    rm.cell(row=_r["n"], column=3).fill = FILL_SUB
    rm.merge_cells(start_row=_r["n"], start_column=2, end_row=_r["n"], end_column=3)
    rm.row_dimensions[_r["n"]].height = 22
    _r["n"] += 1


def bullet(text):
    put("•  " + text, Font(name="Calibri", size=11), valign="top")


def gap():
    _r["n"] += 1


put("BCBA Sourcing Tracker — NC + MD — Read Me",
    Font(name="Calibri", size=18, bold=True, color="1F4E78"), height=28)
put("A sourcing roster built on PUBLIC professional-licensing data.",
    Font(name="Calibri", size=11, italic=True, color="808080"))
gap()

section("1.  The Tabs")
bullet("Candidates — one row per licensee. Type directly into the cells; the green "
       "\"Years Licensed\" column fills in automatically from the issue date.")
bullet("Dashboard — live counts by state, credential, license status, outreach "
       "status, and years-licensed range. Updates as you edit Candidates.")
bullet("Lists — the menu of choices behind every dropdown.")
bullet("Read Me — this tab.")
gap()

section("2.  Where the Data Comes From (public sources)")
bullet("BACB certificant registry — name, credential, certification #, "
       "certification date, status, city/state.")
bullet("NC Behavior Analyst Licensure Board — LBA/LABA license verification.")
bullet("Maryland Board of Professional Counselors & Therapists (BOPC) — "
       "behavior-analyst licensees.")
bullet("LinkedIn URL and Professional Contact are filled in by you during normal "
       "sourcing — look people up one at a time and record the PUBLIC professional "
       "profile / office line / work email only.")
gap()

section("3.  Years Licensed (item b, done properly)")
bullet("Enter the License / Cert Issue Date and the Years Licensed column computes "
       "whole years automatically. Use the Dashboard buckets to target by tenure.")
gap()

section("4.  Fields intentionally NOT in this workbook")
bullet("Home / residential address, age/date of birth, and personal (non-work) "
       "phone or email are deliberately excluded.")
bullet("Age: recording candidates' ages exposes the firm to age-discrimination "
       "(ADEA) claims — it has no legitimate role in sourcing.")
bullet("Home address + personal contact: aggregating these on a roster of named "
       "private individuals is a privacy/safety risk, not a recruiting need.")
bullet("Reach candidates through LinkedIn and published professional channels — "
       "the same way the Delaware pipeline worked.")
gap()

section("5.  Before You Start")
bullet("Rows 2–3 of Candidates are illustrative samples. Delete them before "
       "entering real data.")

wb.active = wb.sheetnames.index("Candidates")

wb.save(OUT)
print("Saved", OUT)
