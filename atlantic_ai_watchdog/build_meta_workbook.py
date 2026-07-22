#!/usr/bin/env python3
"""Build Meta_Plaintiffs.xlsx from the aggregated CSVs (+ video pool if present)."""
import csv, os, re, sys
import openpyxl

ILLEGAL = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f]")
def clean_val(v):
    return ILLEGAL.sub("", v) if isinstance(v, str) else v
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HF = PatternFill("solid", fgColor="1F3864"); HFONT = Font(bold=True, color="FFFFFF")
RED = PatternFill("solid", fgColor="F4CCCC")


def style(ws, n):
    for c in range(1, n + 1):
        ws.cell(1, c).fill = HF; ws.cell(1, c).font = HFONT
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(n)}1"


def add(wb, title, rows, strong_on="total_books"):
    if not rows:
        return
    fld = list(rows[0].keys())
    ws = wb.create_sheet(title); ws.append(fld)
    si = fld.index(strong_on) + 1 if strong_on in fld else None
    for r in rows:
        ws.append([clean_val(r[k]) for k in fld])
        if si:
            try:
                if int(r[strong_on]) >= 10:
                    ws.cell(ws.max_row, si).fill = RED
            except ValueError:
                pass
    style(ws, len(fld))
    for c in range(1, len(fld) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["B"].width = 40


def load(path):
    return list(csv.DictReader(open(path, encoding="utf-8"))) if os.path.exists(path) else []


def main():
    wb = openpyxl.Workbook()
    clean = load("meta_plaintiffs_clean.csv")
    allr = load("meta_plaintiffs.csv")
    video = load("meta_video_channels.csv")

    ci = wb.active; ci.title = "Case Info"; ci.column_dimensions["A"].width = 114
    strong_n = sum(1 for r in clean if int(r["total_books"]) >= 10)
    lines = [
        "META PLAINTIFF POOL — potential clients affected by Meta", "",
        "Basis: Sullivan et al. v. Meta Platforms & Mark Zuckerberg, N.D. Cal. 3:26-cv-06793 (07/02/26).",
        "Meta used pirated books from shadow libraries (LibGen, Books3) to train Llama. Claims: copyright",
        "infringement (direct + contributory) + DMCA 1202. Filed class = textbook authors; this list = ANY author.",
        "Meta also used the HD-VILA-100M video dataset -> YouTube creators are a second Meta pool (Video tab).", "",
        "FILTERS: org/imprint/artifact names removed (author_type); public-domain authors removed via Project",
        "Gutenberg match (public_domain flag). 'Viable' tabs = individual AND not public-domain.",
        "Ranked by total_books = claim strength.", "",
        f"Book authors: {len(allr):,} total names; {len(clean):,} viable; {strong_n:,} with 10+ books.",
        f"Video channels (Meta/HD-VILA): {len(video):,}." if video else
        "Video channels (Meta/HD-VILA): harvest in progress.", "",
        "CAVEAT: filters are heuristic (name-based); verify each name before outreach. LibGen coverage depends",
        "on harvest completeness — see whether this is the SAMPLE or FULL version in the counts above.",
    ]
    for t in lines:
        ci.append([t]); ci.cell(ci.max_row, 1).alignment = Alignment(wrap_text=True)
    ci["A1"].font = Font(bold=True, size=14, color="1F3864")

    add(wb, "Strong Book Authors (10+)", [r for r in clean if int(r["total_books"]) >= 10])
    jewish = load("jewish_book_authors_meta.csv")
    if jewish:
        add(wb, "Jewish Book Authors (Meta)", jewish, strong_on="total_books")
    add(wb, "All Viable Book Authors", clean)
    if video:
        add(wb, "Video Channels (Meta)", video, strong_on="total_books")
    add(wb, "All Names (incl orgs, PD)", allr)

    wb.save(sys.argv[1] if len(sys.argv)>1 else "Meta_Plaintiffs.xlsx")
    print("wrote Meta_Plaintiffs.xlsx | viable book authors:", len(clean),
          "| strong:", strong_n, "| video:", len(video))


if __name__ == "__main__":
    main()
